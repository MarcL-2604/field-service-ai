"""
Mapping-Logik fuer echte SMax-Daten (KI_gestuetzte_Planung.xlsx).

5 Sheets:
  1_Skills Matrix             → SMaxSkillEintrag (JA/NEIN → PM oder None)
  2_Durchscnittliche_Zeit_MC  → SMaxEinsatzDauer (Stunden → Minuten x60)
  3_Closed Jobs               → SMaxGeschlossenAuftrag (historisch_manager_einsatz Flag)
  4_Open_Jobs                 → SMaxOffenerAuftrag (auftragstyp=UNBEKANNT)
  5_Wohnorte                  → SMaxTechniker (nur Germany, Manager ausgeschlossen)

Manager (nicht aktiv dispatchbar, historische Daten bleiben erhalten):
  Stefan Theuerkorn, Rolf Gieling, Juergen Lehmann

Oeffentliche API:
  map_skill_row()          — einzelne Skill-Zeile mappen
  map_einsatzdauer_row()   — einzelne Einsatzdauer-Zeile mappen
  map_closed_job_row()     — einzelne Closed-Job-Zeile mappen
  map_open_job_row()       — einzelne Open-Job-Zeile mappen
  map_wohnort_row()        — einzelne Wohnort-Zeile mappen (None = gefiltert)
  parse_smax_xlsx()        — vollstaendige XLSX-Datei parsen (benoetigt openpyxl)
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Optional

from pydantic import BaseModel

from api.cluster_mapping import finde_cluster
from techniker.plz_lookup import plz_fuer_stadt


# ── Manager-Ausschluss ─────────────────────────────────────────────────────────

MANAGER_NAMEN: frozenset[str] = frozenset({
    "Stefan Theuerkorn",
    "Rolf Gieling",
    "Jürgen Lehmann",
})


def ist_manager(name: str) -> bool:
    """True wenn der Name einem der 3 Manager entspricht."""
    return name.strip() in MANAGER_NAMEN


# ── Pydantic-Modelle fuer SMax-Daten ──────────────────────────────────────────

class SMaxSkillEintrag(BaseModel):
    """Ein Skill-Eintrag aus Sheet 1_Skills Matrix (eine Zelle = ein Eintrag)."""
    model_code: str              # z.B. MC-12345
    tech_name: str               # Vollstaendiger Technikername aus Spaltenheader
    qualifikation: Optional[str] = None  # "PM+Repair", "PM", oder None
    cluster: Optional[str] = None        # z.B. "CLUSTER1_OR", None wenn unbekannt
    repair: Optional[bool] = None        # True wenn Repair-Qualifikation vorhanden


class SMaxEinsatzDauer(BaseModel):
    """Einsatzdauer-Eintrag aus Sheet 2_Durchscnittliche_Zeit_MC."""
    model_code: str
    mittelwert_min: int          # Mittelwert in Minuten (Stunden x 60)
    median_min: int              # Median in Minuten (Stunden x 60)
    bemerkung: Optional[str] = None   # z.B. "inkl. FA"


class SMaxGeschlossenAuftrag(BaseModel):
    """Geschlossener Auftrag aus Sheet 3_Closed Jobs."""
    auftragsnummer: str
    account: str
    ort: str
    plz: Optional[str] = None
    model_code: str
    next_pm_due_date: Optional[str] = None
    erledigung_datum: Optional[str] = None    # Actual Resolution (Datum+Zeit)
    seriennummer: Optional[str] = None
    warranty_date: Optional[str] = None
    techniker: Optional[str] = None
    historisch_manager_einsatz: bool = False  # True wenn Technician = Manager


class SMaxOffenerAuftrag(BaseModel):
    """Offener Auftrag aus Sheet 4_Open_Jobs."""
    auftragsnummer: str
    account: str
    ort: str
    plz: Optional[str] = None
    model_code: str
    next_pm_due_date: Optional[str] = None
    seriennummer: Optional[str] = None
    warranty_date: Optional[str] = None
    contract_end_date: Optional[str] = None
    auftrags_status: str                    # Order Status (Scheduled/On Hold/...)
    on_hold_grund: Optional[str] = None    # On Hold Reason
    auftragstyp: str = "UNBEKANNT"         # nicht in Daten – manuell klaeren


class SMaxTechniker(BaseModel):
    """Techniker aus Sheet 5_Wohnorte (nur Germany, Manager ausgeschlossen)."""
    name: str
    land: str
    ort: str
    plz: Optional[str] = None
    plz_unsicher: bool = False    # True fuer Wildenberg und Linden
    ist_aktiv: bool = True        # False fuer Manager (wird hier nicht benoetigt,
                                  # da Manager beim Import gefiltert werden)
    hugo_ka: bool = False         # manuell zu pflegen – Default immer False


# ── Gesamt-Ergebnis ────────────────────────────────────────────────────────────

@dataclass
class SMaxSheetErgebnis:
    """Ergebnis fuer ein einzelnes Sheet."""
    sheet_name: str
    zeilen_gesamt: int = 0
    zeilen_ok: int = 0
    zeilen_gefiltert: int = 0
    fehler: list[str] = field(default_factory=list)
    warnungen: list[str] = field(default_factory=list)


@dataclass
class SMaxImportErgebnis:
    """Kombiniertes Ergebnis aller 5 Sheets."""
    skills: list[SMaxSkillEintrag] = field(default_factory=list)
    einsatzdauern: list[SMaxEinsatzDauer] = field(default_factory=list)
    geschlossene_auftraege: list[SMaxGeschlossenAuftrag] = field(default_factory=list)
    offene_auftraege: list[SMaxOffenerAuftrag] = field(default_factory=list)
    techniker: list[SMaxTechniker] = field(default_factory=list)
    sheet_ergebnisse: list[SMaxSheetErgebnis] = field(default_factory=list)
    log: list[str] = field(default_factory=list)

    @property
    def aktive_techniker_anzahl(self) -> int:
        return sum(1 for t in self.techniker if t.ist_aktiv)


# ── Zeilen-Mapper (reine Funktionen, testbar ohne XLSX) ───────────────────────

def map_skill_row(model_code: str, tech_name: str, wert: str) -> SMaxSkillEintrag:
    """Mappt eine einzelne Skill-Matrix-Zelle mit Cluster-Mapping.

    JA + cluster + repair=True  → qualifikation="PM+Repair"
    JA + cluster + repair=False → qualifikation="PM"
    JA + kein cluster           → qualifikation="PM", cluster=None
    NEIN                        → qualifikation=None
    """
    mc = str(model_code).strip()
    ist_ja = str(wert).strip().upper() == "JA"

    if not ist_ja:
        return SMaxSkillEintrag(model_code=mc, tech_name=str(tech_name).strip())

    info = finde_cluster(mc)
    if info is None:
        return SMaxSkillEintrag(
            model_code=mc,
            tech_name=str(tech_name).strip(),
            qualifikation="PM",
            cluster=None,
            repair=None,
        )

    qualifikation = "PM+Repair" if info.repair else "PM"
    return SMaxSkillEintrag(
        model_code=mc,
        tech_name=str(tech_name).strip(),
        qualifikation=qualifikation,
        cluster=info.cluster,
        repair=info.repair,
    )


def map_einsatzdauer_row(row: dict) -> SMaxEinsatzDauer:
    """Mappt eine Einsatzdauer-Zeile; rechnet Stunden → Minuten um (x60)."""
    def _stunden_zu_min(v: object) -> int:
        try:
            return round(float(str(v).replace(",", ".")) * 60)
        except (ValueError, TypeError):
            return 0

    bemerkung = str(row.get("Bemerkung", "") or "").strip() or None
    return SMaxEinsatzDauer(
        model_code=str(row.get("Model_Code", "")).strip(),
        mittelwert_min=_stunden_zu_min(row.get("Mittelwert")),
        median_min=_stunden_zu_min(row.get("Median")),
        bemerkung=bemerkung,
    )


def map_closed_job_row(row: dict) -> SMaxGeschlossenAuftrag:
    """Mappt eine Zeile aus 3_Closed Jobs.

    Technician-Feld wird gegen MANAGER_NAMEN geprueft.
    GCH-Code-Filter ist deaktiviert (kein GCH in echten Daten).
    """
    techniker = str(row.get("Technician", "") or "").strip() or None
    return SMaxGeschlossenAuftrag(
        auftragsnummer=str(row.get("Work Order Number", "")).strip(),
        account=str(row.get("Account", "")).strip(),
        ort=str(row.get("City", "")).strip(),
        plz=str(row.get("Zip", "") or "").strip() or None,
        model_code=str(row.get("Model Number", "")).strip(),
        next_pm_due_date=str(row.get("Next PM Due Date", "") or "").strip() or None,
        erledigung_datum=str(row.get("Actual Resolution", "") or "").strip() or None,
        seriennummer=str(row.get("Serial Number", "") or "").strip() or None,
        warranty_date=str(row.get("Warranty End Date", "") or "").strip() or None,
        techniker=techniker,
        historisch_manager_einsatz=ist_manager(techniker) if techniker else False,
    )


def map_open_job_row(row: dict) -> SMaxOffenerAuftrag:
    """Mappt eine Zeile aus 4_Open_Jobs.

    Auftragstyp wird als UNBEKANNT belassen — nicht raten.
    GCH-Code-Filter ist deaktiviert.
    """
    status = str(row.get("Order Status", "") or "").strip()
    on_hold = str(row.get("On Hold Reason", "") or "").strip() or None
    return SMaxOffenerAuftrag(
        auftragsnummer=str(row.get("Work Order Number", "")).strip(),
        account=str(row.get("Account", "")).strip(),
        ort=str(row.get("City", "")).strip(),
        plz=str(row.get("Zip", "") or "").strip() or None,
        model_code=str(row.get("Model Number", "")).strip(),
        next_pm_due_date=str(row.get("Next PM Due Date", "") or "").strip() or None,
        seriennummer=str(row.get("Serial Number", "") or "").strip() or None,
        warranty_date=str(row.get("Warranty End Date", "") or "").strip() or None,
        contract_end_date=str(row.get("Contract End Date", "") or "").strip() or None,
        auftrags_status=status,
        on_hold_grund=on_hold,
        auftragstyp="UNBEKANNT",
    )


def map_wohnort_row(row: dict) -> Optional[SMaxTechniker]:
    """Mappt eine Zeile aus 5_Wohnorte.

    Gibt None zurueck wenn:
    - Land != "Germany"
    - Name ist Manager

    PLZ wird per plz_fuer_stadt() ermittelt.
    hugo_ka bleibt False (manuell zu pflegen).
    """
    land = str(row.get("Land", "") or "").strip()
    if land != "Germany":
        return None

    name = str(row.get("Name", "") or "").strip()
    if ist_manager(name):
        return None

    ort = str(row.get("Stadt", "") or "").strip()
    plz, unsicher = plz_fuer_stadt(ort)

    return SMaxTechniker(
        name=name,
        land=land,
        ort=ort,
        plz=plz,
        plz_unsicher=unsicher,
        ist_aktiv=True,
        hugo_ka=False,
    )


# ── XLSX-Parser (benoetigt openpyxl) ──────────────────────────────────────────

def parse_smax_xlsx(
    workbook_bytes: bytes,
    sample_limit: Optional[int] = None,
) -> SMaxImportErgebnis:
    """Parst alle 5 Sheets einer KI_gestuetzte_Planung.xlsx.

    Args:
        workbook_bytes: XLSX-Datei als Bytes (z.B. aus FastAPI UploadFile).
        sample_limit:   Wenn gesetzt, werden nur die ersten N Daten-Zeilen
                        je Sheet verarbeitet (z.B. 20 fuer Testlaeufe).
                        None = alle Zeilen (Produktion nach Freigabe).

    Returns:
        SMaxImportErgebnis mit allen geparstem Daten und Log-Meldungen.
    """
    import io
    import openpyxl

    ergebnis = SMaxImportErgebnis()

    try:
        wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    except Exception as e:
        ergebnis.log.append(f"FEHLER: XLSX konnte nicht geoeffnet werden: {e}")
        return ergebnis

    # ── Sheet 1: Skills Matrix ─────────────────────────────────────────────────
    sheet_res = SMaxSheetErgebnis(sheet_name="1_Skills Matrix")
    ergebnis.sheet_ergebnisse.append(sheet_res)
    if "1_Skills Matrix" in wb.sheetnames:
        ws = wb["1_Skills Matrix"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            # Erste Zeile = Header: [None/Model_Code, Tech1, Tech2, ...]
            headers = [str(h or "").strip() for h in rows[0]]
            # Manager-Spalten und leere Spalten ueberspringen
            tech_cols = [
                (idx, name)
                for idx, name in enumerate(headers)
                if idx > 0 and name and not ist_manager(name)
            ]
            daten_rows = rows[1:]
            if sample_limit:
                daten_rows = daten_rows[:sample_limit]
            for row in daten_rows:
                mc = str(row[0] or "").strip() if row else ""
                if not mc:
                    sheet_res.zeilen_gefiltert += 1
                    continue
                sheet_res.zeilen_gesamt += 1
                for col_idx, tech_name in tech_cols:
                    if col_idx < len(row):
                        eintrag = map_skill_row(mc, tech_name, row[col_idx] or "")
                        ergebnis.skills.append(eintrag)
                sheet_res.zeilen_ok += 1
    else:
        sheet_res.warnungen.append("Sheet '1_Skills Matrix' nicht gefunden")

    # ── Sheet 2: Einsatzdauer ──────────────────────────────────────────────────
    sheet_res2 = SMaxSheetErgebnis(sheet_name="2_Durchscnittliche_Zeit_MC")
    ergebnis.sheet_ergebnisse.append(sheet_res2)
    if "2_Durchscnittliche_Zeit_MC" in wb.sheetnames:
        ws2 = wb["2_Durchscnittliche_Zeit_MC"]
        rows2 = list(ws2.iter_rows(values_only=True))
        if rows2:
            hdrs2 = [str(h or "").strip() for h in rows2[0]]
            daten2 = rows2[1:]
            if sample_limit:
                daten2 = daten2[:sample_limit]
            for row in daten2:
                sheet_res2.zeilen_gesamt += 1
                row_dict = {hdrs2[i]: row[i] for i in range(min(len(hdrs2), len(row)))}
                try:
                    ergebnis.einsatzdauern.append(map_einsatzdauer_row(row_dict))
                    sheet_res2.zeilen_ok += 1
                except Exception as e:
                    sheet_res2.fehler.append(f"Zeile: {e}")
    else:
        sheet_res2.warnungen.append("Sheet '2_Durchscnittliche_Zeit_MC' nicht gefunden")

    # ── Sheet 3: Closed Jobs ───────────────────────────────────────────────────
    sheet_res3 = SMaxSheetErgebnis(sheet_name="3_Closed Jobs")
    ergebnis.sheet_ergebnisse.append(sheet_res3)
    if "3_Closed Jobs" in wb.sheetnames:
        ws3 = wb["3_Closed Jobs"]
        rows3 = list(ws3.iter_rows(values_only=True))
        if rows3:
            hdrs3 = [str(h or "").strip() for h in rows3[0]]
            daten3 = rows3[1:]
            if sample_limit:
                daten3 = daten3[:sample_limit]
            manager_zaehler = 0
            for row in daten3:
                sheet_res3.zeilen_gesamt += 1
                row_dict = {hdrs3[i]: row[i] for i in range(min(len(hdrs3), len(row)))}
                try:
                    auftrag = map_closed_job_row(row_dict)
                    ergebnis.geschlossene_auftraege.append(auftrag)
                    if auftrag.historisch_manager_einsatz:
                        manager_zaehler += 1
                    sheet_res3.zeilen_ok += 1
                except Exception as e:
                    sheet_res3.fehler.append(f"Zeile: {e}")
            if manager_zaehler:
                sheet_res3.warnungen.append(
                    f"{manager_zaehler} Eintraege mit Manager als Techniker "
                    f"(historisch_manager_einsatz=True, nicht in Statistiken)"
                )
    else:
        sheet_res3.warnungen.append("Sheet '3_Closed Jobs' nicht gefunden")

    # ── Sheet 4: Open Jobs ─────────────────────────────────────────────────────
    sheet_res4 = SMaxSheetErgebnis(sheet_name="4_Open_Jobs")
    ergebnis.sheet_ergebnisse.append(sheet_res4)
    if "4_Open_Jobs" in wb.sheetnames:
        ws4 = wb["4_Open_Jobs"]
        rows4 = list(ws4.iter_rows(values_only=True))
        if rows4:
            hdrs4 = [str(h or "").strip() for h in rows4[0]]
            daten4 = rows4[1:]
            if sample_limit:
                daten4 = daten4[:sample_limit]
            for row in daten4:
                sheet_res4.zeilen_gesamt += 1
                row_dict = {hdrs4[i]: row[i] for i in range(min(len(hdrs4), len(row)))}
                try:
                    ergebnis.offene_auftraege.append(map_open_job_row(row_dict))
                    sheet_res4.zeilen_ok += 1
                except Exception as e:
                    sheet_res4.fehler.append(f"Zeile: {e}")
    else:
        sheet_res4.warnungen.append("Sheet '4_Open_Jobs' nicht gefunden")

    # ── Sheet 5: Wohnorte ──────────────────────────────────────────────────────
    sheet_res5 = SMaxSheetErgebnis(sheet_name="5_Wohnorte")
    ergebnis.sheet_ergebnisse.append(sheet_res5)
    if "5_Wohnorte" in wb.sheetnames:
        ws5 = wb["5_Wohnorte"]
        rows5 = list(ws5.iter_rows(values_only=True))
        if rows5:
            hdrs5 = [str(h or "").strip() for h in rows5[0]]
            for row in rows5[1:]:
                sheet_res5.zeilen_gesamt += 1
                row_dict = {hdrs5[i]: row[i] for i in range(min(len(hdrs5), len(row)))}
                tech = map_wohnort_row(row_dict)
                if tech is None:
                    sheet_res5.zeilen_gefiltert += 1
                else:
                    ergebnis.techniker.append(tech)
                    sheet_res5.zeilen_ok += 1
    else:
        sheet_res5.warnungen.append("Sheet '5_Wohnorte' nicht gefunden")

    # ── Zusammenfassung ────────────────────────────────────────────────────────
    n_aktiv = ergebnis.aktive_techniker_anzahl
    ergebnis.log.append(
        f"{n_aktiv} Techniker aktiv importiert, "
        f"3 Manager ausgeschlossen (Stefan Theuerkorn, Rolf Gieling, Jürgen Lehmann), "
        f"deren Auftragshistorie bleibt erhalten"
    )
    ergebnis.log.append(
        f"Skills: {len(ergebnis.skills)} Eintraege | "
        f"Einsatzdauern: {len(ergebnis.einsatzdauern)} | "
        f"Closed Jobs: {len(ergebnis.geschlossene_auftraege)} | "
        f"Open Jobs: {len(ergebnis.offene_auftraege)} | "
        f"Techniker: {n_aktiv}"
    )

    return ergebnis
